from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
A="Arial"; INK="FF1A1A1A"; RED="FFC00000"; AMB="FFB26B00"; GREY="FF7F7F7F"; GREEN="FF008000"; BLUE="FF1546C9"
HDR=PatternFill("solid",fgColor="FF1F2A44"); BAND=PatternFill("solid",fgColor="FFE8ECF4")
YEL=PatternFill("solid",fgColor="FFFFFF00")
thin=Side(style="thin",color="FFD0D5DE"); BOX=Border(left=thin,right=thin,top=thin,bottom=thin)
def F(s=10,b=False,c=INK,i=False): return Font(name=A,size=s,bold=b,color=c,italic=i)

wb=Workbook(); ws=wb.active; ws.title="Punch List"
ws["B2"]="Build 1 — Punch List"; ws["B2"].font=F(16,True)
ws["B3"]="Every defect found in Day_4_Skeleten.xlsx. Fix in place, top to bottom. Tick the Done column as you go."
ws["B3"].font=F(10,i=True,c=GREY)
ws["B4"]="25 substantive items in 6 groups, plus 28 spelling errors listed at the end."
ws["B4"].font=F(10,True)

ITEMS=[
 ("GROUP A — WRONG NUMBERS. These change the model's output.",None,None,None,None),
 ("A1","CRITICAL","r248:r250, cols H–L",
  "The cash roll-forward feeding interest income is 15 hardcoded numbers, frozen at an earlier state. It has drifted "
  "from the live statement by 0.70 rising to 4.56. Interest income is earned on cash the model does not have.",
  "H248 =H122   H249 =G250   H250 =SUM(H248:H249)   then fill right to L."),
 ("A2","CRITICAL","r210, cols H–L",
  "Depreciation is charged on CLOSING PP&E (=-H206*H211) but the rate was calibrated on OPENING. Overstates "
  "depreciation by 1.85 to 2.79 a year, and makes r210 and r211 mutually dependent - an unintended circular reference.",
  "H210 =-H206*H208   then fill right. Check H210 becomes -39.559 and H211 becomes 320.399."),

 ("GROUP B — THE REVOLVER. Six separate breaks in one chain.",None,None,None,None),
 ("B1","BROKEN","r185, all columns",
  "Repayment =MAX(MIN(0,cash-min),beginning) returns the BEGINNING BALANCE whenever cash exceeds the minimum, and "
  "returns it while drawing too. Wrong in 4 of 6 stress cases.",
  "H185 =MIN(H183, MAX(0, H176-H181))"),
 ("B2","BROKEN","r186, all columns",
  "Ending =SUM(begin,draw,repay) ADDS a repayment that is computed positive, so the balance grows when it should shrink.",
  "H186 =H183+H184-H185"),
 ("B3","BROKEN","r177, cols I–L",
  "Revolver beginning balance is a typed 0 from I onward instead of the prior year's closing. The roll-forward is "
  "severed even before the formulas are fixed.",
  "I177 =H186   then fill right to L."),
 ("B4","BROKEN","r77, cols H–L",
  "The balance sheet revolver line is EMPTY across the whole forecast. A draw would never reach the balance sheet.",
  "H77 =H186   then fill right to L."),
 ("B5","MISSING","Cash flow financing, r116:r119",
  "There is no revolver line in financing, so a draw would never move cash either.",
  "Insert a row: 'Increase/(Decrease) in Revolver'  =H186-G186"),
 ("B6","WIRING","r176, all columns",
  "Feeds off r250 - the stale series from A1 - AND off final cash. The sweep must test cash BEFORE revolver activity, "
  "or once connected it chases its own tail.",
  "Point r176 at opening cash + operating + investing + financing EXCLUDING the revolver."),

 ("GROUP C — LINKS THAT POINT FORWARD. A closing balance must never read the next year.",None,None,None,None),
 ("C1","STRUCTURE","E157",  "Long-term debt closing reads =F155, the NEXT year's opening.","E157 =SUM(E155:E156)"),
 ("C2","STRUCTURE","E167",  "Short-term debt closing reads =F162.","E167 =SUM(E165:E166)"),
 ("C3","STRUCTURE","F167",  "Short-term debt closing reads =G165.","F167 =SUM(F165:F166)"),
 ("C4","STRUCTURE","G167",  "Short-term debt closing reads =H162.","G167 =SUM(G165:G166)"),
 ("C5","STRUCTURE","E236",  "Common equity closing reads =F234.","E236 =SUM(E234:E235)"),
 ("C6","STRUCTURE","F236",  "Common equity closing reads =G234.","F236 =SUM(F234:F235)"),
 ("C7","STRUCTURE","G236",  "Common equity closing reads =H234.","G236 =SUM(G234:G235)"),
 ("!!","WARNING","Read before doing C1–C7",
  "E156, E166 and E235 currently DERIVE the movement as (ending - beginning). If you change the ending to a SUM "
  "while the movement still derives from it, you create a genuine circular reference.",
  "Do it in this order: first type the actual movement (LT -20 each year; ST 0, 0, -5; equity 0, 0, -8), THEN change "
  "the ending row to the SUM."),

 ("GROUP D — FORMULAS THAT CHANGE SHAPE MID-ROW. Right answers, fragile wiring.",None,None,None,None),
 ("D1","STRUCTURE","r18",  "H uses =G82 (balance sheet); I–L use =I155 (debt schedule). Two routes to the same number.","Pick one route and use it in H through L."),
 ("D2","STRUCTURE","r20",  "H uses =G76; I–L use =I165.","Pick one route and use it in H through L."),
 ("D3","STRUCTURE","r27",  "H uses =G88; I–L use =I236.","Pick one route and use it in H through L."),
 ("D4","STRUCTURE","r162", "H uses =H20; I–L use =I165.","Pick one route and use it in H through L."),
 ("D5","STRUCTURE","r165", "H uses =H162 (same column); I–L use =H167 (prior closing). Correct in H only by coincidence.","H165 =G167, matching the pattern in I–L."),
 ("D6","STRUCTURE","r234", "H uses =H27; I–L use =H236.","Pick one route and use it in H through L."),

 ("GROUP E — DEAD CELLS CARRYING WRONG NUMBERS.",None,None,None,None),
 ("E1","WRONG","E208",
  "Opening PP&E is =SUM(E209:E211), which adds capex, depreciation and CLOSING to get 288. Note 4 says 268, and your "
  "own E11 correctly holds 268. Nothing reads it, so it is silently wrong.",
  "E208 =268, or link it to E11."),
 ("E2","WRONG","E238",
  "Opening retained earnings is =SUM(E239:E241) = 297.6. It should be 264.9 closing less 32.7 retained = 232.2.",
  "E238 =232.2, or E241-E239-E240."),

 ("GROUP F — LABELS.",None,None,None,None),
 ("F1","WRONG","r68 and r69",
  "Goodwill and intangibles are transposed. The report has goodwill 112.0 and intangibles 38.0; your r68 'Intangibles' "
  "holds 112 and r69 'Goodwille' holds 38. The total is right so nothing breaks, but the balance sheet states "
  "something untrue.",
  "Swap the two labels, or swap the two values."),
 ("F2","TIDY","28 rows",
  "Spelling errors throughout - see the Spelling tab. None affect a number, all affect how the model reads to anyone "
  "else.",
  "Work down the Spelling tab."),
]

h=7
for i,t in enumerate(["#","Type","Where","What is wrong","The fix","Done"]):
    c=ws.cell(h,2+i); c.value=t; c.font=F(10,True,c="FFFFFFFF"); c.fill=HDR
COL={"CRITICAL":RED,"BROKEN":RED,"MISSING":RED,"WIRING":RED,"WARNING":AMB,"STRUCTURE":GREY,"WRONG":AMB,"TIDY":GREY}
r=h+1
first_item=r
for it in ITEMS:
    if it[1] is None:
        ws.cell(r,2,it[0]).font=F(10,True)
        for c in range(2,8): ws.cell(r,c).fill=BAND; ws.cell(r,c).border=BOX
        r+=1; continue
    n,typ,where,what,fix=it
    ws.cell(r,2,n).font=F(10,True)
    ws.cell(r,3,typ).font=F(9,True,c=COL[typ])
    ws.cell(r,4,where).font=Font(name="Consolas",size=10,bold=True)
    ws.cell(r,5,what).font=F(10); ws.cell(r,5).alignment=Alignment(wrap_text=True,vertical="top")
    ws.cell(r,6,fix).font=Font(name="Consolas",size=9.5,color=GREEN)
    ws.cell(r,6).alignment=Alignment(wrap_text=True,vertical="top")
    ws.cell(r,7).fill=YEL
    for c in range(2,8): ws.cell(r,c).border=BOX
    ws.row_dimensions[r].height=54
    r+=1
last_item=r-1

dv=DataValidation(type="list",formula1='"Y"',allow_blank=True)
ws.add_data_validation(dv); dv.add(f"H{first_item}:H{last_item}")

for w,c in ((2,"A"),(6,"B"),(11,"C"),(15,"D"),(78,"E"),(52,"F"),(7,"G"),(7,"H")):
    ws.column_dimensions[c].width=w
ws.freeze_panes=f"B{h+1}"

# ---------- verify tab ----------
vs=wb.create_sheet("Verify After")
vs["B2"]="Verify After Fixing"; vs["B2"].font=F(15,True)
vs["B3"]="Run all four. The first three have exact expected answers, so you can tell whether the fix landed."
vs["B3"].font=F(10,i=True,c=GREY)
CHECKS=[
 ("1","Hardcode sweep","Select H:L over the whole model. F5 → Special → Constants → OK.",
  "ONLY the assumption rows (6-29) and the debt/equity movement rows should light up. If r248:r250 still select, "
  "A1 is not fixed."),
 ("2","Depreciation lands right","Read H210 and H211 after fixing A2.",
  "H210 = -39.559 and H211 = 320.399. If H210 still reads -41.411 you are still charging on the closing balance."),
 ("3","The two cash series agree","Compare r124 against r250 across H to L.",
  "They must be identical in every column. Any difference means A1 is incomplete."),
 ("4","Stress the revolver","Set r6 (revenue growth) to -0.05 for H through L.",
  "Cash must fall below 60, the revolver must DRAW, r186 must never go negative, and r94 must stay at zero. "
  "If the revolver stays at zero while cash goes below 60, the chain is still broken."),
]
h2=6
for i,t in enumerate(["#","Check","How","What you should see"]):
    c=vs.cell(h2,2+i); c.value=t; c.font=F(10,True,c="FFFFFFFF"); c.fill=HDR
r=h2+1
for n,chk,how,exp in CHECKS:
    vs.cell(r,2,n).font=F(10,True)
    vs.cell(r,3,chk).font=F(10,True)
    vs.cell(r,4,how).font=F(10); vs.cell(r,4).alignment=Alignment(wrap_text=True,vertical="top")
    vs.cell(r,5,exp).font=F(10,c=GREEN); vs.cell(r,5).alignment=Alignment(wrap_text=True,vertical="top")
    for c in range(2,6): vs.cell(r,c).border=BOX
    vs.row_dimensions[r].height=52; r+=1
vs.cell(r+2,2,"The balance check staying at zero proves nothing on its own.").font=F(11,True,c=RED)
vs.cell(r+3,2,"It was zero before these fixes and it will be zero after. That is the whole lesson from build 1.").font=F(10,c=GREY)
for w,c in ((2,"A"),(5,"B"),(26,"C"),(50,"D"),(86,"E")):
    vs.column_dimensions[c].width=w

# ---------- spelling tab ----------
sp=wb.create_sheet("Spelling")
sp["B2"]="Spelling"; sp["B2"].font=F(15,True)
sp["B3"]="None of these change a number. All of them change how the model reads to someone else."
sp["B3"].font=F(10,i=True,c=GREY)
TYPOS=[(8,"SG&A (Adjustted for Inflation)","Adjusted"),(14,"Accounts Recieveable Day","Receivable Days"),
 (18,"Lonng Term Debt Beginning Balance","Long"),(24,"Revolver Continuence Fee","Commitment Fee"),
 (37,"Gorss Profit","Gross Profit"),(41,"Dperectiation","Depreciation"),
 (62,"Account Recieveables","Accounts Receivable"),(69,"Goodwille","Goodwill"),
 (102,"Depreciationg","Depreciation"),(105,"Changes in Account Recieveable","Accounts Receivable"),
 (109,"Net Change in Cash From Operating Activies","Activities"),(112,"Capital Expenditrue","Expenditure"),
 (113,"Net Change in Cash From Investing Activiities","Activities"),(115,"Cashflow From Financing Activies","Activities"),
 (122,"Nett Change in Cash","Net"),(131,"Account Recieveables","Accounts Receivable"),
 (136,"Accounts Recieveable Day","Receivable Days"),(142,"Account Recieveables","Accounts Receivable"),
 (152,"Lonng Term Debt Beginning Balance","Long"),(179,"Revolver Continuence Fee","Commitment Fee"),
 (188,"Reolver Interest Expense","Revolver"),(189,"Revolver Continuence Fee","Commitment Fee"),
 (194,"Reolver Interest Expense","Revolver"),(195,"Revolver Continuence Fee","Commitment Fee"),
 (210,"(Deprectiation)","(Depreciation)"),(222,"SG&A (Adjustted for Inflation)","Adjusted"),
 (241,"Ending Retained Earniings","Earnings"),(248,"Nett Change in Cash","Net")]
h3=6
for i,t in enumerate(["Row","Currently reads","Should be","Done"]):
    c=sp.cell(h3,2+i); c.value=t; c.font=F(10,True,c="FFFFFFFF"); c.fill=HDR
r=h3+1
for rw,cur,fix in TYPOS:
    sp.cell(r,2,f"r{rw}").font=Font(name="Consolas",size=10,bold=True)
    sp.cell(r,3,cur).font=F(10)
    sp.cell(r,4,fix).font=F(10,True,c=GREEN)
    sp.cell(r,5).fill=YEL
    for c in range(2,6): sp.cell(r,c).border=BOX
    r+=1
dv2=DataValidation(type="list",formula1='"Y"',allow_blank=True)
sp.add_data_validation(dv2); dv2.add(f"E{h3+1}:E{r-1}")
for w,c in ((2,"A"),(8,"B"),(48,"C"),(30,"D"),(8,"E")):
    sp.column_dimensions[c].width=w

wb.calculation.fullCalcOnLoad=True
out='/home/user/launchpad/learning/model-practice/Build_1_punch_list.xlsx'
wb.save(out)
subst=sum(1 for it in ITEMS if it[1] is not None and it[0]!="!!")
print(f"saved {out}")
print(f"  substantive items: {subst}   spelling: {len(TYPOS)}   tabs: {wb.sheetnames}")
